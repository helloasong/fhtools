import json
from datetime import datetime

import pandas as pd
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QSplitter,
    QListWidget, QListWidgetItem, QGroupBox, QLabel,
    QPushButton, QTableWidget, QTableWidgetItem,
    QProgressBar, QMessageBox, QFileDialog, QAbstractItemView,
    QHeaderView, QScrollArea, QWidget
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from typing import List

from src.controllers.project_controller import ProjectController
from src.core.cross_binning import (
    CrossBinningAnalyzer,
    CrossBinningFilters,
    CrossBinningResult,
    CrossBinningHeatmapData,
)
from src.ui.widgets.cross_binning_params import CrossBinningParamsPanel
from src.ui.widgets.cross_binning_heatmap import CrossBinningHeatmap


class CrossBinningWorker(QThread):
    """组合策略分析工作线程"""

    finished = pyqtSignal(object)  # CrossBinningResult
    error = pyqtSignal(str)

    def __init__(self, controller: ProjectController, features: List[str], filters: CrossBinningFilters):
        super().__init__()
        self.controller = controller
        self.features = features
        self.filters = filters

    def run(self):
        try:
            df = self.controller.df
            target_col = self.controller.state.target_col
            configs = {
                f: self.controller.state.binning_configs[f]
                for f in self.features
            }

            # 获取各变量过滤后的数据
            filtered_map = {}
            for feat in self.features:
                filtered_map[feat] = self.controller.get_filtered_data_for_feature(feat)

            result = CrossBinningAnalyzer.analyze(
                df=df,
                target_col=target_col,
                features=self.features,
                configs=configs,
                filters=self.filters,
                filtered_data_map=filtered_map,
            )

            self.finished.emit(result)
        except Exception as e:
            self.error.emit(str(e))


class CrossBinningHeatmapWorker(QThread):
    """热力图数据构建工作线程"""

    finished = pyqtSignal(object)  # CrossBinningHeatmapData
    error = pyqtSignal(str)

    def __init__(self, controller: ProjectController, feature_x: str, feature_y: str):
        super().__init__()
        self.controller = controller
        self.feature_x = feature_x
        self.feature_y = feature_y

    def run(self):
        try:
            df = self.controller.df
            target_col = self.controller.state.target_col
            config_x = self.controller.state.binning_configs[self.feature_x]
            config_y = self.controller.state.binning_configs[self.feature_y]

            filtered_map = {
                self.feature_x: self.controller.get_filtered_data_for_feature(self.feature_x),
                self.feature_y: self.controller.get_filtered_data_for_feature(self.feature_y),
            }

            heatmap_data = CrossBinningAnalyzer.build_heatmap_data(
                df=df,
                target_col=target_col,
                feature_x=self.feature_x,
                feature_y=self.feature_y,
                config_x=config_x,
                config_y=config_y,
                filtered_data_map=filtered_map,
            )
            self.finished.emit(heatmap_data)
        except Exception as e:
            self.error.emit(str(e))


class CrossBinningDialog(QDialog):
    """组合策略分析对话框"""

    def __init__(self, controller: ProjectController, parent=None):
        super().__init__(parent)
        self.controller = controller
        self.worker = None
        self.heatmap_worker = None
        self.current_result = None

        self.setWindowTitle("🔀 组合策略分析")
        self.setMinimumSize(1100, 750)
        self.resize(1200, 800)

        self.init_ui()
        self.load_variables()
        self.update_button_state()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(8)

        # 顶部工具栏
        toolbar = QHBoxLayout()
        self.title_label = QLabel("<b>🔀 组合策略分析</b>")
        self.title_label.setStyleSheet("font-size: 16px;")

        self.export_btn = QPushButton("📥 导出规则")
        self.export_btn.setEnabled(False)
        self.export_btn.setToolTip("导出选中的策略规则")
        self.export_btn.clicked.connect(self.on_export)

        self.close_btn = QPushButton("关闭")
        self.close_btn.clicked.connect(self.reject)

        toolbar.addWidget(self.title_label)
        toolbar.addStretch()
        toolbar.addWidget(self.export_btn)
        toolbar.addWidget(self.close_btn)
        layout.addLayout(toolbar)

        # 主分割器
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # 左侧面板
        left_panel = self._build_left_panel()
        splitter.addWidget(left_panel)

        # 右侧面板
        right_panel = self._build_right_panel()
        splitter.addWidget(right_panel)

        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 3)
        splitter.setSizes([280, 900])

        layout.addWidget(splitter)

    def _build_left_panel(self) -> QWidget:
        """构建左侧面板"""
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

        panel = QWidget()
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(4, 4, 4, 4)
        layout.setSpacing(8)

        # 变量选择组
        var_group = QGroupBox("选择变量（勾选已分箱变量）")
        var_layout = QVBoxLayout(var_group)
        var_layout.setSpacing(4)

        self.variable_list = QListWidget()
        self.variable_list.setSelectionMode(QListWidget.SelectionMode.NoSelection)
        self.variable_list.itemChanged.connect(self.on_variable_selection_changed)
        var_layout.addWidget(self.variable_list)

        layout.addWidget(var_group)

        # 参数配置组
        self.params_panel = CrossBinningParamsPanel()
        layout.addWidget(self.params_panel)

        # 预览信息
        self.preview_label = QLabel("请选择至少2个已分箱变量")
        self.preview_label.setStyleSheet(
            "color: #666; padding: 8px; background: #F5F5F5; border-radius: 6px;"
        )
        self.preview_label.setWordWrap(True)
        layout.addWidget(self.preview_label)

        # 分析按钮
        self.analyze_btn = QPushButton("▶️ 开始分析")
        self.analyze_btn.setEnabled(False)
        self.analyze_btn.setStyleSheet(
            """
            QPushButton {
                background: #4CAF50; color: white;
                border-radius: 6px; padding: 10px;
                font-weight: bold; font-size: 14px;
            }
            QPushButton:hover { background: #43A047; }
            QPushButton:disabled { background: #BDBDBD; }
            """
        )
        self.analyze_btn.clicked.connect(self.on_analyze)
        layout.addWidget(self.analyze_btn)

        # 进度条（初始隐藏）
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 0)
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)

        layout.addStretch()
        scroll.setWidget(panel)
        return scroll

    def _build_right_panel(self) -> QWidget:
        """构建右侧面板"""
        panel = QWidget()
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)

        # 热力图容器（仅2变量时显示）
        self.heatmap_container = QWidget()
        heatmap_layout = QVBoxLayout(self.heatmap_container)
        heatmap_layout.setContentsMargins(0, 0, 0, 0)
        self.heatmap_widget = CrossBinningHeatmap()
        heatmap_layout.addWidget(self.heatmap_widget)
        self.heatmap_container.setVisible(False)
        layout.addWidget(self.heatmap_container)

        # 结果表格
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(11)
        self.result_table.setHorizontalHeaderLabels([
            "规则编号", "风险等级", "组合条件", "样本数", "占比",
            "坏样本数", "坏账率", "整体坏账率", "Lift", "WOE", "操作"
        ])
        self.result_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.result_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.result_table.setAlternatingRowColors(True)
        self.result_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.result_table.horizontalHeader().setStretchLastSection(True)
        self.result_table.setMinimumHeight(200)
        layout.addWidget(self.result_table)

        # 状态栏
        self.status_bar = QLabel("就绪 — 请选择变量并点击开始分析")
        self.status_bar.setStyleSheet("color: #666; padding: 4px; background: #FAFAFA; border-radius: 4px;")
        layout.addWidget(self.status_bar)

        return panel

    def load_variables(self):
        """加载变量列表"""
        state = self.controller.state
        if not state or not state.feature_cols:
            return

        binned_features = set(state.binning_configs.keys())

        for feat in state.feature_cols:
            item = QListWidgetItem()
            is_binned = feat in binned_features

            if is_binned:
                n_bins = len(state.binning_configs[feat].splits) - 1
                item.setText(f"{feat}  ({n_bins}箱)")
                item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                item.setCheckState(Qt.CheckState.Unchecked)
                item.setData(Qt.ItemDataRole.UserRole, True)
            else:
                item.setText(f"{feat}  (未分箱)")
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsUserCheckable)
                item.setForeground(Qt.GlobalColor.gray)
                item.setData(Qt.ItemDataRole.UserRole, False)

            self.variable_list.addItem(item)

    def update_button_state(self):
        """根据已分箱变量数更新按钮状态"""
        state = self.controller.state
        if not state:
            return
        binned_count = len(state.binning_configs)
        if binned_count < 2:
            self.analyze_btn.setEnabled(False)
            self.analyze_btn.setToolTip(f"请先对至少2个变量进行分箱 (当前: {binned_count})")
        else:
            self.analyze_btn.setToolTip("")
            # 由 on_variable_selection_changed 根据选择数量控制

    def on_variable_selection_changed(self, item: QListWidgetItem):
        """变量选择变化时更新预览"""
        checked = self._get_checked_features()

        if len(checked) >= 2:
            state = self.controller.state
            bin_counts = [len(state.binning_configs[f].splits) - 1 for f in checked]
            total_combo = 1
            for c in bin_counts:
                total_combo *= c

            preview_lines = [
                "已选变量: " + " × ".join([f"{f}({c}箱)" for f, c in zip(checked, bin_counts)]),
                f"笛卡尔积组合数: {total_combo}",
            ]
            if total_combo > 5000:
                preview_lines.append("⚠️ 组合数超过5000上限，无法分析")
                self.analyze_btn.setEnabled(False)
            else:
                self.analyze_btn.setEnabled(True)

            self.preview_label.setText("\n".join(preview_lines))
            self.preview_label.setStyleSheet(
                "color: #333; padding: 8px; background: #E8F5E9; border-radius: 6px;"
            )
        else:
            need = 2 - len(checked)
            self.preview_label.setText(f"已选 {len(checked)} 个变量，还需选择 {need} 个")
            self.preview_label.setStyleSheet(
                "color: #666; padding: 8px; background: #F5F5F5; border-radius: 6px;"
            )
            self.analyze_btn.setEnabled(False)

    def _get_checked_features(self) -> List[str]:
        """获取当前勾选的变量名"""
        checked = []
        for i in range(self.variable_list.count()):
            item = self.variable_list.item(i)
            if item.checkState() == Qt.CheckState.Checked:
                feat = item.text().split("  ")[0]
                checked.append(feat)
        return checked

    def on_analyze(self):
        """执行分析"""
        features = self._get_checked_features()
        if len(features) < 2:
            QMessageBox.warning(self, "提示", "请至少选择2个变量")
            return

        filters = self.params_panel.get_filters()

        # 显示进度
        self.progress_bar.setVisible(True)
        self.analyze_btn.setEnabled(False)
        self.analyze_btn.setText("分析中...")
        self.export_btn.setEnabled(False)
        self.status_bar.setText("正在计算组合策略...")

        # 启动工作线程
        self.worker = CrossBinningWorker(self.controller, features, filters)
        self.worker.finished.connect(self.on_analyze_finished)
        self.worker.error.connect(self.on_analyze_error)
        self.worker.start()

    def on_analyze_finished(self, result: CrossBinningResult):
        """分析完成回调"""
        self.current_result = result
        self.progress_bar.setVisible(False)
        self.analyze_btn.setEnabled(True)
        self.analyze_btn.setText("▶️ 开始分析")
        self.export_btn.setEnabled(True)

        # 更新状态栏
        filters = self.params_panel.get_filters()
        if filters.show_all:
            self.status_bar.setText(
                f"【不过滤模式】整体坏账率: {result.overall_bad_rate:.2%} | "
                f"总组合: {result.total_combinations} | "
                f"展示全部 {len(result.rules)} 条组合"
            )
            self.status_bar.setStyleSheet(
                "color: #1565C0; padding: 4px; background: #E3F2FD; border-radius: 4px;"
            )
        else:
            self.status_bar.setText(
                f"整体坏账率: {result.overall_bad_rate:.2%} | "
                f"总组合: {result.total_combinations} | "
                f"命中规则: {result.filtered_combinations}"
            )
            self.status_bar.setStyleSheet(
                "color: #666; padding: 4px; background: #FAFAFA; border-radius: 4px;"
            )

        # 2变量时显示热力图
        if len(result.feature_names) == 2:
            self._load_heatmap(result.feature_names[0], result.feature_names[1])
        else:
            self.heatmap_container.setVisible(False)

        # 填充表格
        self._fill_result_table(result)

    def on_analyze_error(self, error_msg: str):
        """分析错误回调"""
        self.progress_bar.setVisible(False)
        self.analyze_btn.setEnabled(True)
        self.analyze_btn.setText("▶️ 开始分析")
        self.status_bar.setText(f"分析失败: {error_msg}")
        self.status_bar.setStyleSheet(
            "color: #D32F2F; padding: 4px; background: #FFEBEE; border-radius: 4px;"
        )
        QMessageBox.critical(self, "分析失败", error_msg)

    def _load_heatmap(self, feature_x: str, feature_y: str):
        """异步加载热力图数据"""
        self.heatmap_worker = CrossBinningHeatmapWorker(
            self.controller, feature_x, feature_y
        )
        self.heatmap_worker.finished.connect(self._on_heatmap_ready)
        self.heatmap_worker.error.connect(lambda e: None)  # 热力图失败静默处理
        self.heatmap_worker.start()

    def _on_heatmap_ready(self, heatmap_data: CrossBinningHeatmapData):
        """热力图数据就绪"""
        self.heatmap_widget.set_data(heatmap_data)
        self.heatmap_container.setVisible(True)

    def _fill_result_table(self, result: CrossBinningResult):
        """填充结果表格"""
        self.result_table.setRowCount(len(result.rules))

        risk_display = {
            "extreme-high": "🔴 极高",
            "high": "🟠 高",
            "normal": "⚪ 正常",
            "low": "🟢 低",
        }

        # 不过滤模式下给高风险/优质客群加颜色标记
        filters = self.params_panel.get_filters()
        show_all = filters.show_all

        for i, rule in enumerate(result.rules):
            values = [
                rule.rule_id,
                risk_display.get(rule.risk_level, rule.risk_level),
                rule.condition_str,
                str(rule.sample_count),
                f"{rule.sample_rate:.2%}",
                str(rule.bad_count),
                f"{rule.bad_rate:.2%}",
                f"{rule.overall_bad_rate:.2%}",
                f"{rule.lift:.2f}",
                f"{rule.woe:.4f}",
                "📋",
            ]
            for j, v in enumerate(values):
                item = QTableWidgetItem(v)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

                # 组合条件列设置 Tooltip 显示完整条件
                if j == 2:
                    item.setToolTip(rule.condition_str)

                # 不过滤模式下加背景色
                if show_all:
                    if rule.risk_level in ("extreme-high", "high"):
                        item.setBackground(Qt.GlobalColor.red)
                        item.setForeground(Qt.GlobalColor.white)
                    elif rule.risk_level == "low":
                        item.setBackground(Qt.GlobalColor.green)
                        item.setForeground(Qt.GlobalColor.black)

                self.result_table.setItem(i, j, item)

        # 数字类列紧凑显示
        for col in [0, 1, 3, 4, 5, 6, 7, 8, 9, 10]:
            self.result_table.horizontalHeader().setSectionResizeMode(col, QHeaderView.ResizeMode.ResizeToContents)
        # 组合条件列拉伸占据剩余空间
        self.result_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        # 设置组合条件列最小宽度，避免被压缩
        self.result_table.setColumnWidth(2, 320)

    def on_export(self):
        """导出规则"""
        if not self.current_result or not self.current_result.rules:
            QMessageBox.information(self, "提示", "没有可导出的规则")
            return

        # 简单导出：让用户选择文件路径，导出 JSON
        path, _ = QFileDialog.getSaveFileName(
            self,
            "导出策略规则",
            "cross_binning_rules.json",
            "JSON (*.json);;Excel (*.xlsx);;文本 (*.txt)"
        )
        if not path:
            return

        try:
            if path.endswith(".json"):
                self._export_json(path)
            elif path.endswith(".xlsx"):
                self._export_excel(path)
            else:
                self._export_text(path)

            QMessageBox.information(self, "导出成功", f"规则已导出到:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e))

    def _export_json(self, path: str):
        """导出为 JSON"""
        rules_data = []
        for r in self.current_result.rules:
            rules_data.append({
                "rule_id": r.rule_id,
                "risk_level": r.risk_level,
                "conditions": [
                    {"variable": c["variable"], "bin_label": c["bin_label"]}
                    for c in r.conditions
                ],
                "bad_rate": round(r.bad_rate, 4),
                "lift": round(r.lift, 2),
                "sample_count": r.sample_count,
                "sample_rate": round(r.sample_rate, 4),
                "woe": round(r.woe, 4),
            })

        data = {
            "overall_bad_rate": round(self.current_result.overall_bad_rate, 4),
            "total_combinations": self.current_result.total_combinations,
            "filtered_combinations": self.current_result.filtered_combinations,
            "features": self.current_result.feature_names,
            "rules": rules_data,
        }

        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def _export_excel(self, path: str):
        """导出为专业格式化的 Excel（多 Sheet + 颜色标记 + 自适应列宽）"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            # fallback: 基础导出
            df = self.current_result.to_dataframe()
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="策略规则", index=False)
            return

        result = self.current_result
        filters = self.params_panel.get_filters()

        wb = Workbook()

        # ===== 通用样式 =====
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        high_risk_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        low_risk_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

        risk_display = {
            "extreme-high": "极高",
            "high": "高",
            "normal": "正常",
            "low": "低",
        }

        # ===== Sheet 1: 策略规则 =====
        ws = wb.active
        ws.title = "策略规则"
        headers = ["规则编号", "风险等级", "组合条件", "样本数", "占比",
                   "坏样本数", "坏账率", "整体坏账率", "Lift", "WOE", "IV"]

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row_idx, rule in enumerate(result.rules, 2):
            row_fill = None
            if rule.risk_level in ("extreme-high", "high"):
                row_fill = high_risk_fill
            elif rule.risk_level == "low":
                row_fill = low_risk_fill

            row_data = [
                rule.rule_id,
                risk_display.get(rule.risk_level, rule.risk_level),
                rule.condition_str,
                rule.sample_count,
                rule.sample_rate,
                rule.bad_count,
                rule.bad_rate,
                rule.overall_bad_rate,
                rule.lift,
                rule.woe,
                rule.iv,
            ]

            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = left_align if col_idx == 3 else cell_align

                if row_fill:
                    cell.fill = row_fill

                # 数值格式
                if col_idx == 5:       # 占比
                    cell.number_format = "0.00%"
                elif col_idx in (7, 8):  # 坏账率
                    cell.number_format = "0.00%"
                elif col_idx == 9:     # Lift
                    cell.number_format = "0.00"
                elif col_idx == 10:    # WOE
                    cell.number_format = "0.0000"
                elif col_idx == 11:    # IV
                    cell.number_format = "0.0000"

        col_widths = [12, 10, 55, 10, 10, 10, 10, 12, 10, 10, 10]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

        # ===== Sheet 2: 统计摘要 =====
        ws2 = wb.create_sheet("统计摘要")
        summary = [
            ("指标", "值"),
            ("分析时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("整体坏账率", result.overall_bad_rate),
            ("总组合数", result.total_combinations),
            ("命中规则数", result.filtered_combinations),
            ("导出规则数", len(result.rules)),
            ("参与变量", "、".join(result.feature_names)),
        ]
        for feat, n in result.feature_bin_counts.items():
            summary.append((f"  {feat} 箱数", n))

        for r_idx, (k, v) in enumerate(summary, 1):
            c1 = ws2.cell(row=r_idx, column=1, value=k)
            c2 = ws2.cell(row=r_idx, column=2, value=v)
            c1.border = thin_border
            c2.border = thin_border
            if r_idx == 1:
                c1.font = header_font
                c1.fill = header_fill
                c2.font = header_font
                c2.fill = header_fill
            c1.alignment = left_align
            c2.alignment = cell_align
            if isinstance(v, float):
                c2.number_format = "0.00%"

        ws2.column_dimensions["A"].width = 22
        ws2.column_dimensions["B"].width = 45

        # ===== Sheet 3: 筛选参数 =====
        ws3 = wb.create_sheet("筛选参数")
        sort_map = {"bad_rate_desc": "坏账率降序", "lift_desc": "Lift降序", "sample_desc": "样本数降序"}
        params = [
            ("参数名", "值"),
            ("最小样本占比", f"{filters.min_sample_rate:.2%}"),
            ("高风险倍数 (≥)", filters.bad_rate_high_multiplier),
            ("优质客群倍数 (≤)", filters.bad_rate_low_multiplier),
            ("最小 Lift", filters.min_lift),
            ("排序方式", sort_map.get(filters.sort_by, filters.sort_by)),
            ("展示全部组合", "是" if filters.show_all else "否"),
        ]
        for r_idx, (k, v) in enumerate(params, 1):
            c1 = ws3.cell(row=r_idx, column=1, value=k)
            c2 = ws3.cell(row=r_idx, column=2, value=v)
            c1.border = thin_border
            c2.border = thin_border
            if r_idx == 1:
                c1.font = header_font
                c1.fill = header_fill
                c2.font = header_font
                c2.fill = header_fill
            c1.alignment = left_align
            c2.alignment = cell_align

        ws3.column_dimensions["A"].width = 22
        ws3.column_dimensions["B"].width = 30

        wb.save(path)

    def _export_text(self, path: str):
        """导出为文本规则"""
        lines = [
            "# 组合策略规则",
            f"# 整体坏账率: {self.current_result.overall_bad_rate:.2%}",
            f"# 总组合数: {self.current_result.total_combinations}",
            f"# 命中规则数: {self.current_result.filtered_combinations}",
            "",
        ]
        for r in self.current_result.rules:
            lines.append(
                f"IF {r.condition_str} THEN {r.risk_level} "
                f"(bad_rate={r.bad_rate:.2%}, lift={r.lift:.2f}, n={r.sample_count})"
            )

        with open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))

    def closeEvent(self, event):
        """关闭时停止工作线程"""
        if self.worker and self.worker.isRunning():
            self.worker.quit()
            self.worker.wait(1000)
        if self.heatmap_worker and self.heatmap_worker.isRunning():
            self.heatmap_worker.quit()
            self.heatmap_worker.wait(1000)
        event.accept()
