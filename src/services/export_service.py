import os
from typing import Dict, List, Optional
import pandas as pd
import numpy as np

from src.data.models import (
    ProjectState, FilterRule, FilterCondition, FilterLogicNode,
    FilterMode, FeatureFilterSetting,
)
from src.utils.formatting import format_bin_label, resolve_precision_step


def _format_filter_rule(state: ProjectState, feature: str) -> Optional[str]:
    """格式化特征的过滤规则为可读文本

    Returns:
        过滤规则说明文本，无过滤时返回 None
    """
    setting = state.feature_filter_settings.get(feature)
    global_rule = state.global_filter_rule

    # 确定生效的规则
    effective_rule = None
    rule_source = None

    if setting:
        if setting.mode == FilterMode.DISABLED:
            return "过滤规则: 不应用任何过滤（使用全部样本）"
        elif setting.mode == FilterMode.CUSTOM:
            effective_rule = setting.rule
            rule_source = "自定义过滤规则"
        elif setting.mode == FilterMode.GLOBAL:
            effective_rule = global_rule
            rule_source = "全局过滤规则"
    else:
        # 未配置时使用全局规则
        effective_rule = global_rule
        rule_source = "全局过滤规则"

    if not effective_rule or not effective_rule.enabled or not effective_rule.root:
        return None

    def _format_node(node, indent: str = "") -> str:
        if isinstance(node, FilterCondition):
            prefix = "NOT " if node.negate else ""
            if node.operator in ('is null', 'is not null'):
                return f"{indent}{prefix}{node.variable} {node.operator}"
            if node.operator == 'between' and isinstance(node.value, (list, tuple)) and len(node.value) == 2:
                return f"{indent}{prefix}{node.variable} {node.operator} [{node.value[0]}, {node.value[1]}]"
            if isinstance(node.value, list):
                values = ', '.join(str(v) for v in node.value)
                return f"{indent}{prefix}{node.variable} {node.operator} [{values}]"
            return f"{indent}{prefix}{node.variable} {node.operator} {node.value}"

        elif isinstance(node, FilterLogicNode):
            lines = [f"{indent}[{node.operator}]"]
            for child in node.children:
                lines.append(_format_node(child, indent + "  "))
            return '\n'.join(lines)
        return ""

    rule_text = _format_node(effective_rule.root)
    return f"过滤规则 ({rule_source}):\n{rule_text}"


def _build_summary_df(state: ProjectState) -> pd.DataFrame:
    rows: List[pd.DataFrame] = []
    for feature, metrics in state.binning_results.items():
        cfg = state.binning_configs.get(feature)
        precision = (resolve_precision_step(cfg.params) if cfg else "auto")
        df = metrics.summary_table.reset_index().copy()
        df["feature"] = feature
        # 添加过滤规则列
        filter_text = _format_filter_rule(state, feature)
        df["过滤规则"] = filter_text.replace('\n', ' | ') if filter_text else "无"
        # 追加每特征的合计行
        try:
            total = int(df['total'].sum())
            bad_sum = int(df['bad'].sum())
            good_sum = int(df['good'].sum())
            overall_bad_rate = bad_sum / max(total, 1)
            total_row = {
                'feature': feature,
                'bin': '合计',
                'total': total,
                'total_pct': 1.0,
                'bad': bad_sum,
                'bad_rate': overall_bad_rate,
                'good': good_sum,
                'good_dist': np.nan,
                'bad_dist': np.nan,
                'woe': '',
                'iv': '',
                'lift': '',
                '过滤规则': filter_text.replace('\n', ' | ') if filter_text else "无"
            }
            # 保持列顺序并避免分类类型冲突
            df = pd.concat([df, pd.DataFrame([total_row])[df.columns]], ignore_index=True)
        except Exception:
            pass
        if 'bin' in df.columns:
            df['bin'] = df['bin'].apply(lambda b: format_bin_label(b, precision=precision))
        # 统一列顺序（若列存在）
        cols = [c for c in [
            "feature", "bin", "total", "total_pct", "bad", "bad_rate",
            "good", "good_dist", "bad_dist", "woe", "iv", "lift", "过滤规则"
        ] if c in df.columns]
        df = df[cols]
        rows.append(df)
    if not rows:
        return pd.DataFrame()
    return pd.concat(rows, ignore_index=True)


def export_excel(state: ProjectState, output_dir: str) -> str:
    os.makedirs(output_dir, exist_ok=True)
    out_path = os.path.join(output_dir, "binning_report.xlsx")

    summary_df = _build_summary_df(state)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        if not summary_df.empty:
            summary_df.to_excel(writer, index=False, sheet_name="Summary")
        # 每个特征一张表
        for feature, metrics in state.binning_results.items():
            cfg = state.binning_configs.get(feature)
            precision = (resolve_precision_step(cfg.params) if cfg else "auto")
            df = metrics.summary_table.reset_index()
            try:
                total = int(df['total'].sum())
                bad_sum = int(df['bad'].sum())
                good_sum = int(df['good'].sum())
                overall_bad_rate = bad_sum / max(total, 1)
                overall_good_rate = 1.0 - overall_bad_rate
                total_row = {
                    'bin': '合计',
                    'total': total,
                    'total_pct': 1.0,
                    'bad': bad_sum,
                    'bad_rate': overall_bad_rate,
                    'good': good_sum,
                    'good_dist': np.nan,
                    'bad_dist': np.nan,
                    'woe': '',
                    'iv': '',
                    'lift': ''
                }
                # 仅添加存在的列，保持列顺序
                for col in list(df.columns):
                    if col not in total_row:
                        total_row[col] = ''
                df = pd.concat([df, pd.DataFrame([total_row])[df.columns]], ignore_index=True)
            except Exception:
                pass
            if 'bin' in df.columns:
                df['bin'] = df['bin'].apply(lambda b: format_bin_label(b, precision=precision))
            df.to_excel(writer, index=False, sheet_name=str(feature)[:31])

    # 使用 openpyxl 为每个 sheet 添加过滤规则注释和可见说明
    try:
        from openpyxl import load_workbook
        from openpyxl.comments import Comment
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = load_workbook(out_path)
        for feature in state.binning_results:
            rule_text = _format_filter_rule(state, feature)
            if rule_text:
                sheet_name = str(feature)[:31]
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    # 在 A1 单元格添加注释
                    ws['A1'].comment = Comment(rule_text, "FHBinningTool")
                    # 在表头下方插入一行可见的过滤条件说明
                    ws.insert_rows(2)
                    cell = ws.cell(row=2, column=1)
                    cell.value = rule_text.replace('\n', ' | ')
                    cell.font = Font(color="666666", italic=True, size=10)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    # 合并该行所有列
                    max_col = ws.max_column
                    if max_col > 1:
                        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
        wb.save(out_path)
        wb.close()
    except Exception:
        # 注释添加失败不影响主导出功能
        pass

    return out_path


def export_python(state: ProjectState, output_dir: str) -> str:
    os.makedirs(output_dir, exist_ok=True)
    out_path = os.path.join(output_dir, "scorecard_transform.py")

    lines: List[str] = []
    lines.append("import pandas as pd")
    lines.append("")
    lines.append("def transform(df: pd.DataFrame) -> pd.DataFrame:")
    lines.append("    df = df.copy()")

    for feature, metrics in state.binning_results.items():
        # 生成分箱切点与 WOE 映射
        splits = state.binning_configs[feature].splits
        df_metrics = metrics.summary_table.reset_index()
        # 将区间转字符串作为键
        woe_map = {str(row["bin"]): float(row["woe"]) for _, row in df_metrics.iterrows()}

        lines.append(f"    # {feature}")
        lines.append(
            f"    bins_{feature} = {repr(splits)}"
        )
        lines.append(
            f"    cats_{feature} = pd.cut(df[{repr(feature)}], bins=bins_{feature}, include_lowest=True).astype(str)"
        )
        ms = state.binning_configs[feature].missing_strategy
        merge_label = state.binning_configs[feature].missing_merge_label
        if ms == 'separate':
            lines.append(f"    cats_{feature} = cats_{feature}.fillna('Missing')")
        elif ms == 'merge' and merge_label is not None:
            lines.append(f"    cats_{feature} = cats_{feature}.fillna({repr(merge_label)})")
        lines.append(
            f"    woe_map_{feature} = {repr(woe_map)}"
        )
        lines.append(
            f"    df['woe_{feature}'] = cats_{feature}.map(woe_map_{feature})"
        )
        lines.append("")

    lines.append("    return df")

    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return out_path


def export_sql(state: ProjectState, output_dir: str) -> str:
    os.makedirs(output_dir, exist_ok=True)
    out_path = os.path.join(output_dir, "scorecard_rules.sql")
    lines: List[str] = []
    for feature, cfg in state.binning_configs.items():
        lines.append(f"-- Rules for {feature}")
        lines.append(f"CASE")
        splits = cfg.splits
        # Build intervals
        for i in range(len(splits) - 1):
            left = splits[i]
            right = splits[i + 1]
            conds = []
            if np.isfinite(left):
                conds.append(f"{feature} > {left}")
            else:
                conds.append(f"{feature} IS NOT NULL")
            if np.isfinite(right):
                conds.append(f"{feature} <= {right}")
            rule_cond = " AND ".join(conds)
            # find WOE for this bin
            dfm = state.binning_results[feature].summary_table.reset_index()
            woe_val = None
            for _, row in dfm.iterrows():
                if str(row['bin']).find(str(left)) != -1 and str(row['bin']).find(str(right)) != -1:
                    woe_val = float(row['woe'])
                    break
            if woe_val is None:
                woe_val = 0.0
            lines.append(f"  WHEN ({rule_cond}) THEN {woe_val}")
        # Missing strategy
        if cfg.missing_strategy == 'separate':
            lines.append(f"  WHEN {feature} IS NULL THEN /* Missing */ 0.0")
        elif cfg.missing_strategy == 'merge' and cfg.missing_merge_label:
            lines.append(f"  WHEN {feature} IS NULL THEN /* Merge to {cfg.missing_merge_label} */ 0.0")
        lines.append("END AS woe_" + feature)
        lines.append("")

    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return out_path
